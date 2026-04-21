import io

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.models.department import Department


def _seed_employee(
    client: TestClient,
    auth_headers: dict[str, str],
    department: Department,
    **overrides: object,
) -> None:
    base: dict[str, object] = {
        "first_name": "Ada",
        "last_name": "Lovelace",
        "email": "ada@example.com",
        "title": "Engineer",
        "department_id": department.id,
        "hire_date": "2024-01-15",
        "is_active": True,
    }
    base.update(overrides)
    r = client.post("/api/employees", json=base, headers=auth_headers)
    assert r.status_code == 201


def test_analyze_csv_happy_path(client: TestClient, auth_headers: dict[str, str]) -> None:
    csv_bytes = b"name,email,title\nAda,ada@x.com,Engineer\nGrace,grace@x.com,Manager\n"
    r = client.post(
        "/api/reports/analyze-csv",
        files={"file": ("roster.csv", io.BytesIO(csv_bytes), "text/csv")},
        headers=auth_headers,
    )
    assert r.status_code == 200
    body = r.json()
    assert body["filename"] == "roster.csv"
    assert body["rows"] == 2
    assert body["columns"] == 3
    assert body["headers"] == ["name", "email", "title"]
    assert body["preview"][0] == ["Ada", "ada@x.com", "Engineer"]


def test_analyze_rejects_non_csv(client: TestClient, auth_headers: dict[str, str]) -> None:
    r = client.post(
        "/api/reports/analyze-csv",
        files={"file": ("roster.txt", io.BytesIO(b"hello"), "text/plain")},
        headers=auth_headers,
    )
    assert r.status_code == 400


def test_analyze_unauthenticated_returns_401(client: TestClient) -> None:
    r = client.post(
        "/api/reports/analyze-csv",
        files={"file": ("roster.csv", io.BytesIO(b"a,b\n1,2"), "text/csv")},
    )
    assert r.status_code == 401


def test_employees_csv_export(
    client: TestClient, auth_headers: dict[str, str], department: Department
) -> None:
    _seed_employee(client, auth_headers, department)
    r = client.get("/api/reports/employees.csv", headers=auth_headers)
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("text/csv")
    assert "attachment" in r.headers["content-disposition"].lower()
    text = r.content.decode("utf-8-sig")
    assert "first_name" in text
    assert "Ada" in text
    assert "Lovelace" in text


def test_employees_xlsx_export(
    client: TestClient, auth_headers: dict[str, str], department: Department
) -> None:
    _seed_employee(
        client,
        auth_headers,
        department,
        first_name="Grace",
        last_name="Hopper",
        email="grace@example.com",
    )
    r = client.get("/api/reports/employees.xlsx", headers=auth_headers)
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("application/vnd.openxmlformats")

    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["Employees"]
    header_row = [c.value for c in ws[1]]
    assert "first_name" in header_row
    first_data_row = [c.value for c in ws[2]]
    assert "Grace" in first_data_row
    assert "Hopper" in first_data_row


def test_employees_csv_unauthenticated_returns_401(client: TestClient) -> None:
    r = client.get("/api/reports/employees.csv")
    assert r.status_code == 401


def test_employees_xlsx_unauthenticated_returns_401(client: TestClient) -> None:
    r = client.get("/api/reports/employees.xlsx")
    assert r.status_code == 401
